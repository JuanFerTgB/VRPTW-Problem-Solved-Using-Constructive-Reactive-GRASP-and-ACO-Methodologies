import openpyxl
from openpyxl.styles import NamedStyle

# Crear estilo para los porcentajes
percentage_style = NamedStyle(name="percentage_style", number_format="0.00%")

# Datos completos que serán ingresados en el archivo Excel
data = [
    ['VRPTW1.txt', 264.79999999999995, 98.99, 167.50, 4, 3, 33.33, 83],
    ['VRPTW2.txt', 411.92599999999993, 141.82, 190.45, 4, 1, 300.00, 99],
    ['VRPTW3.txt', 640.183, 276.78, 131.30, 9, 2, 350.00, 80],
    ['VRPTW4.txt', 643.5390000000001, 276.78, 132.51, 8, 1, 700.00, 85],
    ['VRPTW5.txt', 620.4760000000001, 178.41, 247.79, 7, 3, 133.33, 83],
    ['VRPTW6.txt', 592.623, 178.41, 232.18, 6, 1, 500.00, 90],
    ['VRPTW7.txt', 601.69, 195.66, 207.52, 11, 5, 120.00, 350],
    ['VRPTW8.txt', 947.36, 264.11, 258.71, 13, 2, 550.00, 366],
    ['VRPTW9.txt', 1248.073, 417.38, 199.02, 17, 4, 325.00, 314],
    ['VRPTW10.txt', 1282.08, 417.38, 207.17, 18, 1, 1700.00, 319],
    ['VRPTW11.txt', 1359.559, 315.74, 330.60, 14, 5, 180.00, 327],
    ['VRPTW12.txt', 1432.4559999999997, 315.74, 353.69, 15, 1, 1400.00, 331],
    ['VRPTW13.txt', 2272.7560000000003, 417.30, 444.63, 30, 10, 200.00, 1264],
    ['VRPTW14.txt', 2342.8950000000004, 492.47, 375.74, 30, 3, 900.00, 1342],
    ['VRPTW15.txt', 2418.648, 562.25, 330.17, 37, 8, 362.50, 1292],
    ['VRPTW16.txt', 1913.6399999999999, 562.25, 240.35, 29, 2, 1350.00, 1329],
    ['VRPTW17.txt', 2774.6209999999996, 564.00, 391.96, 30, 9, 233.33, 1286],
    ['VRPTW18.txt', 2547.3669999999997, 564.00, 351.66, 30, 2, 1400.00, 1317],
]



# Crear un archivo Excel
wb = openpyxl.Workbook()
ws = wb.active

# Definir las columnas
columns = ["Filename", "Total Distance", "Lower Bound Distance", "GAP Distance (%)", 
           "Actual Routes", "Lower Bound Routes", "GAP Routes (%)", "Execution Time (ms)"]
ws.append(columns)

# Agregar los datos fila por fila
for row in data:
    ws.append(row)

# Aplicar formato de porcentaje a las columnas correspondientes
for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):  # Columna de GAP Distance (%)
    for cell in row:
        cell.value = cell.value / 100  # Convertir el valor a decimal

for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):  # Columna de GAP Routes (%)
    for cell in row:
        cell.value = cell.value / 100  # Convertir el valor a decimal

# Guardar el archivo Excel
output_path = r'D:\UNIVERSIDAD\OCTAVO SEMESTRE\prueba 1\VRPTW_Results.xlsx'
wb.save(output_path)

